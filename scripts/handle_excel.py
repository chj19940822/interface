# -*- coding: utf-8 -*-
"""
------------------------------------
@Time    : 2020/6/28 13:00
@Author  : Destiny
@FileName: handle_excel.py
@Software: PyCharm
------------------------------------
"""
from openpyxl import load_workbook
from scripts.handle_config import do_config
from scripts.constants import Datas_File_Path


class HandleExcel:
    """
    定义excel操作类
    """

    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        """
        获取测试数据
        :return:
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            wd = wb.active
        else:
            wd = wb[self.sheetname]
        head_data = tuple(wd.iter_rows(max_row=1, values_only=True))[0]
        list = []
        for datas in tuple(wd.iter_rows(min_row=2, values_only=True)):
            list.append(dict(zip(head_data, datas)))
        return list

    def write_result(self, row, actual, result):
        """
        在指定行写入数据
        :param row:行号
        :param actual:实际结果
        :param result:用例执行的结果(Pass或者Fail)
        :return:
        """
        other_wb = load_workbook(self.filename)
        if self.sheetname is None:
            other_wd = other_wb.active
        else:
            other_wd = other_wb[self.sheetname]
        if isinstance(row, int) and (2 <= row <= other_wd.max_row):
            other_wd.cell(row=row, column=do_config.get_int("excel", "actual_col"), value=actual)
            other_wd.cell(row=row, column=do_config.get_int("excel", "result_col"), value=result)
            other_wb.save(self.filename)
            other_wb.close()
        else:
            print("传入的行号有误, 行号应为大于1的整数")



if __name__ == '__main__':
    filename = Datas_File_Path
    do_excel = HandleExcel(filename,she)
    cases = do_excel.get_cases()
    print(cases)
    do_excel.write_result(2,"出现", "测试")
    pass
